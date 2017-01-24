#-*- coding: utf-8 -*-
from openpyxl import load_workbook
import os


# wb = load_workbook(filename=r'existing_file.xlsx')
# sheets = wb.get_sheet_names()  # 获取所有表格(worksheet)的名字
# sheet0 = sheets[0]  # 第一个表格的名称
# ws = wb.get_sheet_by_name('sheet_name')  # 获取特定的 worksheet
#
# # 获取表格所有行和列，两者都是可迭代的
# rows = ws.rows
# columns = ws.columns
#
# # 行迭代
# content = []
# for row in rows:
#     line = [col.value for col in row]
#     content.append(line)
#
# # 通过坐标读取值
# print ws.cell('B12').value  # B 表示列，12 表示行
# print ws.cell(row=12, column=2).value
os.chdir('/home/zxjd/Web/cace')
def xlsx2str(data):
    name = 'a.xlsx'
    xls_file = open(name, 'wb')
    xls_file.write(data)
    xls_file.close()
    s = ''
    wb = load_workbook(filename=name)
    sheets = wb.get_sheet_names()
    for i in sheets:
        ws = wb.get_sheet_by_name(i)
        rows= ws.rows
        columns = ws.columns
        for row in rows:
            for col in row:
                if isinstance(col.value, unicode):
                    s = s + col.value.encode('utf-8')
    return s
